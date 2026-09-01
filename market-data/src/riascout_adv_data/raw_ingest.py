"""Lossless raw-table ingestion for official SEC CSV and spreadsheet artifacts."""

import codecs
import csv
import hashlib
import json
import re
import shutil
import tempfile
from dataclasses import dataclass
from datetime import UTC, date, datetime
from pathlib import Path, PurePosixPath
from typing import Any
from zipfile import ZipFile

from duckdb import DuckDBPyConnection
from openpyxl import load_workbook

from riascout_adv_data.official_db import OfficialDatabase

NORMALIZED_ROW_NUMBER = "__adv_source_row_number"
NORMALIZED_SOURCE_WIDTH = "__adv_source_width"
NORMALIZED_RAW_VALUES = "__adv_raw_values_json"


class RawIngestError(ValueError):
    """Raised when an official artifact cannot be loaded losslessly."""


@dataclass(frozen=True)
class RawTableRef:
    """Inventory metadata for one source member loaded as a DuckDB table."""

    artifact_id: str
    member_name: str
    table_name: str
    row_count: int
    columns: tuple[str, ...]
    header_row_number: int
    source_encoding: str


@dataclass(frozen=True)
class _Artifact:
    artifact_id: str
    dataset_kind: str
    payload_path: Path


class RawIngestor:
    """Load immutable official artifacts into source-shaped DuckDB tables."""

    def __init__(self, database: OfficialDatabase) -> None:
        """Initialize the ingestor for one official-data database."""
        self._database = database

    def ingest_artifact(self, artifact_id: str) -> tuple[RawTableRef, ...]:
        """Load every CSV/XLSX member atomically and return its raw table inventory."""
        existing = self._existing_inventory(artifact_id)
        if existing:
            return existing
        artifact = self._artifact(artifact_id)
        if not artifact.payload_path.is_file():
            raise FileNotFoundError(artifact.payload_path)

        with tempfile.TemporaryDirectory(prefix="adv-raw-ingest-") as directory:
            members = _materialize_members(artifact.payload_path, Path(directory))
            if not members:
                raise RawIngestError(f"Artifact {artifact_id!r} contains no CSV or XLSX data members")
            created: list[RawTableRef] = []
            with self._database.transaction() as connection:
                for member_name, member_path in members:
                    ref = self._load_member(
                        connection,
                        artifact=artifact,
                        member_name=member_name,
                        member_path=member_path,
                        workspace=Path(directory),
                    )
                    if ref.row_count == 0 and _member_requires_rows(artifact.dataset_kind, member_name):
                        raise RawIngestError(f"Required source member {member_name!r} has no data rows")
                    connection.execute(
                        """
                        INSERT INTO raw_table_inventory (
                            artifact_id, member_name, raw_table_name, row_count,
                            columns_json, header_row_number, source_encoding, ingested_at
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        [
                            ref.artifact_id,
                            ref.member_name,
                            ref.table_name,
                            ref.row_count,
                            json.dumps(ref.columns),
                            ref.header_row_number,
                            ref.source_encoding,
                            datetime.now(UTC),
                        ],
                    )
                    created.append(ref)
                connection.execute(
                    "UPDATE source_artifacts SET ingest_status = 'ingested' WHERE artifact_id = ?",
                    [artifact_id],
                )
        return tuple(created)

    def _artifact(self, artifact_id: str) -> _Artifact:
        with self._database.connection() as connection:
            row = connection.execute(
                "SELECT artifact_id, dataset_kind, payload_path FROM source_artifacts WHERE artifact_id = ?",
                [artifact_id],
            ).fetchone()
        if row is None:
            raise LookupError(f"Official artifact {artifact_id!r} was not recorded")
        return _Artifact(artifact_id=str(row[0]), dataset_kind=str(row[1]), payload_path=Path(str(row[2])))

    def _existing_inventory(self, artifact_id: str) -> tuple[RawTableRef, ...]:
        with self._database.connection() as connection:
            rows = connection.execute(
                """
                SELECT artifact_id, member_name, raw_table_name, row_count, columns_json, header_row_number
                       , source_encoding
                FROM raw_table_inventory
                WHERE artifact_id = ?
                ORDER BY member_name
                """,
                [artifact_id],
            ).fetchall()
        return tuple(
            RawTableRef(
                artifact_id=str(row[0]),
                member_name=str(row[1]),
                table_name=str(row[2]),
                row_count=int(row[3]),
                columns=tuple(json.loads(str(row[4]))),
                header_row_number=int(row[5]),
                source_encoding=str(row[6] or "utf-8"),
            )
            for row in rows
        )

    def _load_member(
        self,
        connection: DuckDBPyConnection,
        *,
        artifact: _Artifact,
        member_name: str,
        member_path: Path,
        workspace: Path,
    ) -> RawTableRef:
        if member_path.suffix.lower() == ".xlsx":
            source_csv_path, _, header_row_number = _xlsx_to_csv(member_path, workspace)
            source_encoding = "xlsx"
            working_encoding = "utf-8"
        else:
            source_csv_path = member_path
            source_encoding = _csv_encoding(member_path)
            working_encoding = source_encoding
            header_row_number = 1
        csv_path, headers = _normalized_csv(
            source_csv_path,
            workspace=workspace,
            source_encoding=working_encoding,
        )
        _validate_headers(headers, member_name=member_name)
        table_name = _raw_table_name(artifact.artifact_id, member_name)
        table_identifier = quote_ident(table_name)
        csv_literal = _quote_literal(str(csv_path))
        artifact_literal = _quote_literal(artifact.artifact_id)
        member_literal = _quote_literal(member_name)
        connection.execute(
            f"""
            CREATE TABLE {table_identifier} AS
            SELECT * EXCLUDE (
                       {quote_ident(NORMALIZED_ROW_NUMBER)},
                       {quote_ident(NORMALIZED_SOURCE_WIDTH)},
                       {quote_ident(NORMALIZED_RAW_VALUES)}
                   ),
                   {artifact_literal}::VARCHAR AS _artifact_id,
                   {member_literal}::VARCHAR AS _source_member,
                   {quote_ident(NORMALIZED_ROW_NUMBER)}::UBIGINT AS _source_row_number,
                   {quote_ident(NORMALIZED_SOURCE_WIDTH)}::INTEGER AS _source_column_count,
                   nullif({quote_ident(NORMALIZED_RAW_VALUES)}, '')::JSON AS _raw_values_json
            FROM read_csv(
                {csv_literal},
                header = true,
                delim = ',',
                quote = '"',
                escape = '"',
                all_varchar = true,
                ignore_errors = false,
                strict_mode = true,
                null_padding = false,
                comment = '',
                encoding = 'utf-8'
            )
            """
        )
        connection.execute(
            f"""
            INSERT INTO raw_row_errors (
                artifact_id, member_name, source_row_number, error_code,
                error_message, raw_values_json, recorded_at
            )
            SELECT _artifact_id, _source_member, _source_row_number,
                   'source_column_count_mismatch',
                   'Source row has ' || _source_column_count ||
                       ' columns; expected {len(headers)}',
                   _raw_values_json, ?
            FROM {table_identifier}
            WHERE _source_column_count <> ?
            """,
            [datetime.now(UTC), len(headers)],
        )
        count_row = connection.execute(f"SELECT count(*) FROM {table_identifier}").fetchone()
        if count_row is None:
            raise RawIngestError(f"Could not count raw table {table_name!r}")
        row_count = int(count_row[0])
        return RawTableRef(
            artifact_id=artifact.artifact_id,
            member_name=member_name,
            table_name=table_name,
            row_count=row_count,
            columns=tuple(headers),
            header_row_number=header_row_number,
            source_encoding=source_encoding,
        )


def quote_ident(value: str) -> str:
    """Quote a DuckDB identifier without allowing SQL syntax injection."""
    return '"' + value.replace('"', '""') + '"'


def _quote_literal(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"


def _materialize_members(payload_path: Path, workspace: Path) -> list[tuple[str, Path]]:
    suffix = payload_path.suffix.lower()
    if suffix in {".csv", ".xlsx"}:
        return [(payload_path.name, payload_path)]
    if suffix != ".zip":
        raise RawIngestError(f"Unsupported official artifact type: {payload_path.suffix}")
    materialized: list[tuple[str, Path]] = []
    with ZipFile(payload_path) as archive:
        for info in archive.infolist():
            member = PurePosixPath(info.filename)
            if member.is_absolute() or ".." in member.parts:
                raise RawIngestError(f"Artifact contains unsafe ZIP member {info.filename!r}")
            if info.is_dir() or member.suffix.lower() not in {".csv", ".xlsx"}:
                continue
            digest = hashlib.sha256(info.filename.encode()).hexdigest()[:12]
            destination = workspace / f"{digest}-{member.name}"
            with archive.open(info) as source, destination.open("wb") as target:
                shutil.copyfileobj(source, target, length=1024 * 1024)
            materialized.append((info.filename, destination))
    return materialized


def _csv_encoding(path: Path) -> str:
    """Detect supported SEC CSV encodings without replacing undecodable bytes."""
    for encoding in ("utf-8", "windows-1252"):
        if _decodes_strictly(path, encoding=encoding):
            return encoding
    return "latin-1"


def _decodes_strictly(path: Path, *, encoding: str) -> bool:
    try:
        decoder = codecs.getincrementaldecoder(encoding)()
        with path.open("rb") as stream:
            while chunk := stream.read(1024 * 1024):
                decoder.decode(chunk)
        decoder.decode(b"", final=True)
    except UnicodeDecodeError:
        return False
    return True


def _normalized_csv(
    path: Path,
    *,
    workspace: Path,
    source_encoding: str,
) -> tuple[Path, list[str]]:
    """Create a strict UTF-8 working CSV while preserving malformed source rows as JSON."""
    destination = workspace / f"{path.stem}-normalized-utf8.csv"
    python_encoding = "utf-8-sig" if source_encoding == "utf-8" else source_encoding
    with (
        path.open(encoding=python_encoding, newline="") as source,
        destination.open("w", encoding="utf-8", newline="") as target_stream,
    ):
        reader = csv.reader(source, strict=False)
        writer = csv.writer(target_stream, lineterminator="\n")
        try:
            headers = [str(value) for value in next(reader)]
        except StopIteration as error:
            raise RawIngestError(f"CSV source {path.name!r} is empty") from error
        writer.writerow([*headers, NORMALIZED_ROW_NUMBER, NORMALIZED_SOURCE_WIDTH, NORMALIZED_RAW_VALUES])
        expected_width = len(headers)
        for source_row_number, row in enumerate(reader, start=1):
            source_width = len(row)
            raw_values = json.dumps(row, ensure_ascii=False) if source_width != expected_width else ""
            fitted = [*row[:expected_width], *([""] * max(0, expected_width - source_width))]
            writer.writerow([*fitted, source_row_number, source_width, raw_values])
    return destination, headers


def _csv_headers(path: Path, *, encoding: str) -> list[str]:
    python_encoding = "utf-8-sig" if encoding == "utf-8" else encoding
    with path.open(newline="", encoding=python_encoding) as stream:
        try:
            row = next(csv.reader(stream))
        except StopIteration as error:
            raise RawIngestError(f"CSV source {path.name!r} is empty") from error
    return [str(value) for value in row]


def _xlsx_to_csv(path: Path, workspace: Path) -> tuple[Path, list[str], int]:
    headers, header_row_number = _xlsx_headers(path)
    destination = workspace / f"{path.stem}-normalized.csv"
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if not workbook.worksheets:
            raise RawIngestError(f"Workbook {path.name!r} has no worksheets")
        sheet = workbook.worksheets[0]
        with destination.open("w", newline="", encoding="utf-8") as stream:
            writer = csv.writer(stream)
            writer.writerow(headers)
            for row_number, cells in enumerate(sheet.iter_rows(values_only=True), start=1):
                if row_number <= header_row_number:
                    continue
                values = [_cell_text(value) for value in cells[: len(headers)]]
                if any(values):
                    writer.writerow(values + [""] * (len(headers) - len(values)))
    finally:
        workbook.close()
    return destination, headers, header_row_number


def _xlsx_headers(path: Path) -> tuple[list[str], int]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if not workbook.worksheets:
            raise RawIngestError(f"Workbook {path.name!r} has no worksheets")
        sheet = workbook.worksheets[0]
        for row_number, cells in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row_number > 50:
                break
            headers = [_cell_text(value) for value in cells]
            normalized = {_normalize_header(value) for value in headers if value}
            has_crd = bool(normalized & {"crd", "crdnumber", "organizationcrdnumber"})
            has_name = bool(normalized & {"name", "legalname", "primarybusinessname"})
            if has_crd and has_name:
                return _name_blank_headers(headers), row_number
    finally:
        workbook.close()
    raise RawIngestError(f"Workbook {path.name!r} has no unique CRD/business-name header row in its first 50 rows")


def _name_blank_headers(headers: list[str]) -> list[str]:
    last_nonempty = max((index for index, value in enumerate(headers) if value), default=-1)
    return [value or f"__unnamed_{index + 1}" for index, value in enumerate(headers[: last_nonempty + 1])]


def _validate_headers(headers: list[str], *, member_name: str) -> None:
    if not headers:
        raise RawIngestError(f"Source member {member_name!r} has no columns")
    normalized: dict[str, str] = {}
    for header in headers:
        key = _normalize_header(header)
        if not key:
            raise RawIngestError(f"Source member {member_name!r} has a blank column name")
        if key in normalized:
            raise RawIngestError(
                f"Source member {member_name!r} has duplicate normalized columns {normalized[key]!r} and {header!r}"
            )
        normalized[key] = header


def _normalize_header(value: str) -> str:
    return re.sub(r"[^a-z0-9]", "", value.lower())


def _raw_table_name(artifact_id: str, member_name: str) -> str:
    stem = re.sub(r"[^a-z0-9]+", "_", Path(member_name).stem.lower()).strip("_") or "source"
    digest = hashlib.sha256(f"{artifact_id}\0{member_name}".encode()).hexdigest()[:12]
    return f"raw_{stem[:48]}_{digest}"


def _member_requires_rows(dataset_kind: str, member_name: str) -> bool:
    normalized = Path(member_name).name.lower()
    if dataset_kind in {"ria_report", "era_report"}:
        return True
    if dataset_kind == "adv_part1":
        return normalized.startswith("ia_adv_base_a")
    if dataset_kind == "advw":
        return "advw" in normalized and "base" in normalized
    return False


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


__all__ = ["RawIngestError", "RawIngestor", "RawTableRef", "quote_ident"]
